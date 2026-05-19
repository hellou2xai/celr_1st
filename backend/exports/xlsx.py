"""
Excel-export helpers. Every export endpoint follows the same shape:

    rows: list[dict] -> xlsx bytes with a styled header row

Uses openpyxl directly (already in requirements.txt). Workbook output is
streamed back as application/vnd.openxmlformats-officedocument...
"""
from __future__ import annotations

import io
from typing import Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(bold=True, color="FFFFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="FF1F2937")
HEADER_ALIGN = Alignment(horizontal="left", vertical="center")


def rows_to_xlsx(rows: list[dict], sheet_name: str = "Export",
                 columns: Optional[list[str]] = None,
                 currency_cols: Iterable[str] = (),
                 percent_cols: Iterable[str] = (),
                 freeze_header: bool = True) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "Export"

    if not rows:
        ws.append(["No data"])
        return _save_to_bytes(wb)

    cols = columns or list(rows[0].keys())
    ws.append(cols)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    cur = set(currency_cols)
    pct = set(percent_cols)

    for r in rows:
        ws.append([r.get(c) for c in cols])

    # Apply column number formats
    for idx, name in enumerate(cols, start=1):
        col_letter = get_column_letter(idx)
        if name in cur:
            for cell in ws[col_letter][1:]:
                cell.number_format = '"$"#,##0.00'
        elif name in pct:
            for cell in ws[col_letter][1:]:
                cell.number_format = '0.0"%"'
        # Auto-ish width
        max_len = max((len(str(r.get(name, ""))) for r in rows), default=0)
        ws.column_dimensions[col_letter].width = min(40, max(10, max_len + 2))

    if freeze_header:
        ws.freeze_panes = "A2"

    return _save_to_bytes(wb)


def _save_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


XLSX_MEDIA = ("application/vnd.openxmlformats-officedocument."
              "spreadsheetml.sheet")


def xlsx_headers(filename: str) -> dict:
    return {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Cache-Control": "no-store",
    }
