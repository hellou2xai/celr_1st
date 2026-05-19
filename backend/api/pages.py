"""
JSON endpoints powering the React pages.

Each endpoint maps to a page in the React frontend. Logic is implemented
inline here for now; over time the heavier queries should move into
backend/analytics.py alongside the existing MCP-facing helpers.

Where data isn't available in the demo Postgres yet (RIP programs,
invoice line matches), endpoints return well-formed empty shapes so the
frontend renders correctly with "no data" states.
"""
from __future__ import annotations

import csv
import io
from datetime import datetime, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Response

from .. import analytics
from ..auth import current_user
from ..db import fetch, fetchrow

router = APIRouter(prefix="/api", tags=["pages"])


def _user(request_user=Depends(current_user)):
    """Dependency that 401s anonymous callers on demand."""
    return request_user


# --------------------------------------------------------------------- #
# Dashboard
# --------------------------------------------------------------------- #

@router.get("/dashboard")
async def dashboard(_=Depends(_user)):
    """One-shot KPI feed for /."""
    # Open PO totals
    po = await fetchrow("""
        SELECT
            COUNT(DISTINCT po.id)               AS open_po_count,
            COUNT(poe.id)                       AS open_line_count,
            COALESCE(SUM((poe.quantity_ordered - poe.quantity_received) * poe.price), 0)
                                                AS open_po_value
        FROM purchase_order po
        JOIN purchase_order_entry poe ON poe.purchase_order_id = po.id
        WHERE po.status < 5
          AND poe.quantity_ordered > poe.quantity_received
    """)

    # Stockouts in last 30 days
    so = await fetchrow("""
        SELECT COUNT(*) AS n
        FROM item i
        WHERE i.inactive = 0
          AND (i.quantity - i.quantity_committed) <= 0
          AND EXISTS (
            SELECT 1 FROM transaction_entry te
            WHERE te.item_id = i.id AND te.quantity > 0
              AND te.transaction_time >= NOW() - INTERVAL '30 days'
          )
    """)

    # Risk rollup of open POs
    rollup_rows = await fetch("""
        WITH po_lines AS (
          SELECT poe.id, poe.item_id,
                 (poe.quantity_ordered - poe.quantity_received) AS qty_open,
                 (poe.quantity_ordered - poe.quantity_received) * poe.price AS line_value,
                 i.quantity AS on_hand, i.quantity_committed AS committed
          FROM purchase_order po
          JOIN purchase_order_entry poe ON poe.purchase_order_id = po.id
          JOIN item i ON poe.item_id = i.id
          WHERE po.status < 5 AND poe.quantity_ordered > poe.quantity_received
        ),
        velocity AS (
          SELECT te.item_id,
                 SUM(te.quantity)/180.0 AS daily_units
          FROM transaction_entry te
          WHERE te.transaction_time >= NOW() - INTERVAL '180 days'
            AND te.quantity > 0
          GROUP BY te.item_id
        )
        SELECT
          CASE
            WHEN COALESCE(v.daily_units, 0) <= 0 THEN 'Dead'
            WHEN (pl.on_hand - pl.committed + pl.qty_open) / NULLIF(v.daily_units, 0) > 365 THEN 'Excess'
            WHEN (pl.on_hand - pl.committed)                / NULLIF(v.daily_units, 0) <= 7 THEN 'Critical'
            WHEN (pl.on_hand - pl.committed)                / NULLIF(v.daily_units, 0) <= 21 THEN 'High'
            WHEN (pl.on_hand - pl.committed)                / NULLIF(v.daily_units, 0) <= 60 THEN 'Moderate'
            ELSE 'Healthy'
          END AS risk,
          COUNT(*)::int AS lines,
          SUM(pl.line_value) AS value
        FROM po_lines pl
        LEFT JOIN velocity v ON v.item_id = pl.item_id
        GROUP BY 1
    """)
    risk_rollup = {r["risk"]: {"lines": int(r["lines"] or 0),
                               "value": float(r["value"] or 0)} for r in rollup_rows}

    # Cancel/reduce — items in Excess + Dead PO lines
    cancel = await fetchrow("""
        WITH po_lines AS (
          SELECT poe.id, poe.item_id,
                 (poe.quantity_ordered - poe.quantity_received) * poe.price AS line_value,
                 i.quantity, i.quantity_committed
          FROM purchase_order po
          JOIN purchase_order_entry poe ON poe.purchase_order_id = po.id
          JOIN item i ON poe.item_id = i.id
          WHERE po.status < 5 AND poe.quantity_ordered > poe.quantity_received
        ),
        velocity AS (
          SELECT te.item_id, SUM(te.quantity)/180.0 AS daily_units
          FROM transaction_entry te
          WHERE te.transaction_time >= NOW() - INTERVAL '180 days' AND te.quantity > 0
          GROUP BY te.item_id
        )
        SELECT COUNT(*)::int AS lines, COALESCE(SUM(pl.line_value),0) AS value
        FROM po_lines pl
        LEFT JOIN velocity v ON v.item_id = pl.item_id
        WHERE COALESCE(v.daily_units, 0) = 0
           OR (pl.quantity - pl.quantity_committed) / NULLIF(v.daily_units,0) > 180
    """)

    # RTV — items with positive on-hand that haven't moved in 180+ days.
    rtv = await fetchrow("""
        SELECT
          COALESCE(SUM(CASE WHEN i.last_received >= NOW() - INTERVAL '90 days'
                            THEN i.quantity * i.cost ELSE 0 END), 0) AS in_window_value,
          COALESCE(SUM(i.quantity * i.cost), 0) AS total_value
        FROM item i
        WHERE i.inactive = 0 AND i.quantity > 0
          AND (i.last_sold IS NULL OR i.last_sold < NOW() - INTERVAL '180 days')
    """)

    # Top suppliers by recoverable cancel/reduce value
    top_supp = await fetch("""
        WITH po_lines AS (
          SELECT po.supplier_id, poe.item_id,
                 (poe.quantity_ordered - poe.quantity_received) * poe.price AS line_value,
                 i.quantity, i.quantity_committed
          FROM purchase_order po
          JOIN purchase_order_entry poe ON poe.purchase_order_id = po.id
          JOIN item i ON poe.item_id = i.id
          WHERE po.status < 5 AND poe.quantity_ordered > poe.quantity_received
        ),
        velocity AS (
          SELECT te.item_id, SUM(te.quantity)/180.0 AS daily_units
          FROM transaction_entry te
          WHERE te.transaction_time >= NOW() - INTERVAL '180 days' AND te.quantity > 0
          GROUP BY te.item_id
        ),
        flagged AS (
          SELECT pl.*, COALESCE(v.daily_units,0) AS daily,
                 CASE
                   WHEN COALESCE(v.daily_units,0) = 0 THEN 'CANCEL'
                   WHEN (pl.quantity - pl.quantity_committed) / NULLIF(v.daily_units,0) > 365 THEN 'CANCEL'
                   WHEN (pl.quantity - pl.quantity_committed) / NULLIF(v.daily_units,0) > 180 THEN 'REDUCE'
                   ELSE 'KEEP'
                 END AS action
          FROM po_lines pl LEFT JOIN velocity v ON v.item_id = pl.item_id
        )
        SELECT s.supplier_name                                AS "SupplierName",
               COUNT(*) FILTER (WHERE action='CANCEL')::int   AS "CancelLines",
               COUNT(*) FILTER (WHERE action='REDUCE')::int   AS "ReduceLines",
               COALESCE(SUM(CASE WHEN action IN ('CANCEL','REDUCE') THEN line_value END), 0)
                                                              AS "RecoverableValue"
        FROM flagged f
        LEFT JOIN supplier s ON s.id = f.supplier_id
        WHERE action IN ('CANCEL','REDUCE')
        GROUP BY s.supplier_name
        ORDER BY "RecoverableValue" DESC NULLS LAST
        LIMIT 8
    """)

    return {
        "open_po_count":    int((po or {}).get("open_po_count") or 0),
        "open_line_count":  int((po or {}).get("open_line_count") or 0),
        "open_po_value":    float((po or {}).get("open_po_value") or 0),
        "cancel_value":     float((cancel or {}).get("value") or 0),
        "cancel_lines":     int((cancel or {}).get("lines") or 0),
        "rtv_value":        float((rtv or {}).get("total_value") or 0),
        "rtv_in_window_value": float((rtv or {}).get("in_window_value") or 0),
        "recent_stockouts": int((so or {}).get("n") or 0),
        "risk_rollup":      risk_rollup,
        "top_cancel_suppliers": top_supp,
    }


# --------------------------------------------------------------------- #
# Open POs / Cancel / Reduce
# --------------------------------------------------------------------- #

_OPEN_POS_BASE_SQL = """
WITH po_lines AS (
  SELECT
    po.id AS po_id,
    po.po_number, po.date_created, po.supplier_id,
    poe.item_id, poe.id AS poe_id,
    poe.quantity_ordered, poe.quantity_received,
    (poe.quantity_ordered - poe.quantity_received) AS qty_open,
    poe.price AS unit_cost,
    (poe.quantity_ordered - poe.quantity_received) * poe.price AS line_value
  FROM purchase_order po
  JOIN purchase_order_entry poe ON poe.purchase_order_id = po.id
  WHERE po.status < 5
    AND poe.quantity_ordered > poe.quantity_received
    AND po.date_created >= NOW() - make_interval(days => $1)
),
velocity AS (
  SELECT te.item_id,
         SUM(te.quantity) AS units_180d,
         SUM(te.quantity)/180.0 AS daily_units
  FROM transaction_entry te
  WHERE te.transaction_time >= NOW() - INTERVAL '180 days'
    AND te.quantity > 0
  GROUP BY te.item_id
)
SELECT
  pl.po_number      AS "PONumber",
  pl.date_created   AS "PODate",
  s.supplier_name   AS "SupplierName",
  i.item_lookup_code AS "UPC",
  i.description     AS "Description",
  d.name            AS "Department",
  pl.quantity_ordered  AS "QtyOrdered",
  pl.quantity_received AS "QtyReceived",
  pl.qty_open       AS "QtyOpen",
  pl.unit_cost      AS "UnitCost",
  pl.line_value     AS "LineValue",
  i.quantity        AS "OnHand",
  (COALESCE(v.units_180d,0)/6.0) AS "AvgMonthlySales",
  CASE WHEN COALESCE(v.daily_units,0) > 0
       THEN (i.quantity - i.quantity_committed)/v.daily_units/30.0 END AS "CurrentMoS",
  CASE WHEN COALESCE(v.daily_units,0) > 0
       THEN (i.quantity - i.quantity_committed + pl.qty_open)/v.daily_units/30.0 END AS "ProjectedMoS",
  CASE
    WHEN COALESCE(v.daily_units,0) = 0 THEN 'Dead'
    WHEN (i.quantity - i.quantity_committed + pl.qty_open) / NULLIF(v.daily_units,0) > 365 THEN 'Excess'
    WHEN (i.quantity - i.quantity_committed)               / NULLIF(v.daily_units,0) <= 7  THEN 'Critical'
    WHEN (i.quantity - i.quantity_committed)               / NULLIF(v.daily_units,0) <= 21 THEN 'High'
    WHEN (i.quantity - i.quantity_committed)               / NULLIF(v.daily_units,0) <= 60 THEN 'Moderate'
    ELSE 'Healthy'
  END AS "Risk",
  CASE
    WHEN COALESCE(v.daily_units,0) = 0 THEN 'CANCEL'
    WHEN (i.quantity - i.quantity_committed) / NULLIF(v.daily_units,0) > 365 THEN 'CANCEL'
    WHEN (i.quantity - i.quantity_committed) / NULLIF(v.daily_units,0) > 180 THEN 'REDUCE'
    ELSE 'KEEP'
  END AS "Action",
  CASE
    WHEN COALESCE(v.daily_units,0) = 0
      THEN 'Dead SKU — never sold'
    WHEN (i.quantity - i.quantity_committed) / NULLIF(v.daily_units,0) > 365
      THEN 'Already 1+ year of cover'
    WHEN (i.quantity - i.quantity_committed) / NULLIF(v.daily_units,0) > 180
      THEN '6+ months of cover; reduce qty'
    ELSE ''
  END AS "Reason"
FROM po_lines pl
JOIN item i ON pl.item_id = i.id
LEFT JOIN department d ON i.department_id = d.id
LEFT JOIN supplier s ON pl.supplier_id = s.id
LEFT JOIN velocity v ON v.item_id = pl.item_id
"""


async def _open_pos_lines(params: dict) -> list[dict]:
    days = int(params.get("days") or 28)
    sql = _OPEN_POS_BASE_SQL
    conds: list[str] = []
    args: list = [days]
    if params.get("supplier"):
        args.append(f"%{params['supplier']}%")
        conds.append(f"s.supplier_name ILIKE ${len(args)}")
    if params.get("product"):
        args.append(f"%{params['product']}%")
        conds.append(f"(i.item_lookup_code ILIKE ${len(args)} OR i.description ILIKE ${len(args)})")
    if params.get("risk") and params["risk"] not in ("", "Any"):
        # Risk filter must be applied as outer query because it's a CASE.
        sql = f"SELECT * FROM ({sql}) q WHERE q.\"Risk\" = ${len(args)+1}"
        args.append(params["risk"])
    if params.get("action") and params["action"] not in ("", "Any"):
        if params["action"] == "NEEDS_ACTION":
            sql = f"SELECT * FROM ({sql}) q2 WHERE q2.\"Action\" IN ('CANCEL','REDUCE')"
        else:
            sql = f"SELECT * FROM ({sql}) q2 WHERE q2.\"Action\" = ${len(args)+1}"
            args.append(params["action"])
    if conds:
        sql = sql.replace("FROM po_lines pl", "FROM po_lines pl") + "\nWHERE " + " AND ".join(conds)
    sql += "\nORDER BY \"LineValue\" DESC NULLS LAST\nLIMIT " + str(int(params.get("limit") or 500))
    return await fetch(sql, *args)


@router.get("/open-pos")
async def open_pos(days: int = 28, supplier: Optional[str] = None,
                   product: Optional[str] = None,
                   risk: Optional[str] = None, action: Optional[str] = None,
                   limit: int = 500, _=Depends(_user)):
    rows = await _open_pos_lines({"days": days, "supplier": supplier,
                                    "product": product, "risk": risk,
                                    "action": action, "limit": limit})
    line_count = len(rows)
    total = sum(float(r["LineValue"] or 0) for r in rows)
    cancel_v = sum(float(r["LineValue"] or 0) for r in rows if r["Action"] == "CANCEL")
    reduce_v = sum(float(r["LineValue"] or 0) for r in rows if r["Action"] == "REDUCE")

    by_supplier: dict[str, dict] = {}
    for r in rows:
        s = r["SupplierName"] or "(none)"
        b = by_supplier.setdefault(s, {"SupplierName": s, "Lines": 0, "Value": 0,
                                       "CancelLines": 0, "ReduceLines": 0,
                                       "RecoverableValue": 0})
        b["Lines"] += 1
        b["Value"] += float(r["LineValue"] or 0)
        if r["Action"] == "CANCEL":
            b["CancelLines"] += 1
            b["RecoverableValue"] += float(r["LineValue"] or 0)
        elif r["Action"] == "REDUCE":
            b["ReduceLines"] += 1
            b["RecoverableValue"] += float(r["LineValue"] or 0)

    by_supplier_list = sorted(by_supplier.values(),
                               key=lambda x: x["RecoverableValue"], reverse=True)
    return {
        "summary": {"line_count": line_count, "total_value": total,
                    "cancel_value": cancel_v, "reduce_value": reduce_v,
                    "recoverable": cancel_v + reduce_v},
        "by_supplier": by_supplier_list,
        "lines": rows,
    }


@router.get("/open-pos/export.csv")
async def open_pos_csv(days: int = 28, supplier: Optional[str] = None,
                       product: Optional[str] = None, risk: Optional[str] = None,
                       action: Optional[str] = None, limit: int = 10000,
                       _=Depends(_user)):
    rows = await _open_pos_lines({"days": days, "supplier": supplier,
                                  "product": product, "risk": risk,
                                  "action": action, "limit": limit})
    buf = io.StringIO()
    if rows:
        w = csv.DictWriter(buf, fieldnames=list(rows[0].keys()))
        w.writeheader()
        for r in rows:
            w.writerow({k: ("" if v is None else v) for k, v in r.items()})
    return Response(content=buf.getvalue(), media_type="text/csv",
                    headers={"Content-Disposition": 'attachment; filename="open-pos.csv"'})


# --------------------------------------------------------------------- #
# RTV
# --------------------------------------------------------------------- #

@router.get("/rtv")
async def rtv(in_window: Optional[str] = None, supplier: Optional[str] = None,
              limit: int = 500, _=Depends(_user)):
    args: list = []
    conds = ["i.inactive = 0", "i.quantity > 0",
             "(i.last_sold IS NULL OR i.last_sold < NOW() - INTERVAL '180 days')"]
    if supplier:
        args.append(f"%{supplier}%")
        conds.append(f"s.supplier_name ILIKE ${len(args)}")
    if in_window == "1":
        conds.append("i.last_received >= NOW() - INTERVAL '90 days'")

    sql = f"""
      SELECT
        i.item_lookup_code AS "UPC",
        i.description AS "Description",
        s.supplier_name AS "SupplierName",
        i.quantity AS "OnHand",
        i.cost AS "Cost",
        (i.quantity * i.cost) AS "InventoryValue",
        i.last_received AS "LastReceived",
        (CURRENT_DATE - i.last_received::date)::int AS "DaysInStore",
        (i.last_received >= NOW() - INTERVAL '90 days') AS "InReturnWindow"
      FROM item i
      LEFT JOIN supplier s ON i.supplier_id = s.id
      WHERE {" AND ".join(conds)}
      ORDER BY (i.quantity * i.cost) DESC NULLS LAST
      LIMIT {int(limit)}
    """
    rows = await fetch(sql, *args)
    total = sum(float(r["InventoryValue"] or 0) for r in rows)
    in_win = sum(float(r["InventoryValue"] or 0) for r in rows if r["InReturnWindow"])
    return {"summary": {"line_count": len(rows), "total_value": total,
                        "in_window_value": in_win}, "rows": rows}


@router.get("/rtv/export.csv")
async def rtv_csv(in_window: Optional[str] = None, supplier: Optional[str] = None,
                  limit: int = 10000, _=Depends(_user)):
    data = await rtv(in_window=in_window, supplier=supplier, limit=limit)
    rows = data["rows"]
    buf = io.StringIO()
    if rows:
        w = csv.DictWriter(buf, fieldnames=list(rows[0].keys()))
        w.writeheader()
        for r in rows:
            w.writerow({k: ("" if v is None else v) for k, v in r.items()})
    return Response(content=buf.getvalue(), media_type="text/csv",
                    headers={"Content-Disposition": 'attachment; filename="rtv.csv"'})


# --------------------------------------------------------------------- #
# Stockouts
# --------------------------------------------------------------------- #

@router.get("/stockouts")
async def stockouts(velocity_months: int = 6, supplier: Optional[str] = None,
                    dept: Optional[str] = None, min_lost: float = 0,
                    with_open_po: Optional[str] = None, limit: int = 500,
                    _=Depends(_user)):
    args: list = [velocity_months]
    conds = ["i.inactive = 0", "(i.quantity - i.quantity_committed) <= 0"]
    if supplier:
        args.append(f"%{supplier}%")
        conds.append(f"s.supplier_name ILIKE ${len(args)}")
    if dept:
        args.append(f"%{dept}%")
        conds.append(f"d.name ILIKE ${len(args)}")

    sql = f"""
      WITH velocity AS (
        SELECT te.item_id,
               SUM(te.quantity) / GREATEST(1, $1 * 30.0) AS daily_units,
               SUM(te.quantity) / NULLIF($1, 0)          AS monthly_units
        FROM transaction_entry te
        WHERE te.transaction_time >= NOW() - make_interval(months => $1)
          AND te.quantity > 0
        GROUP BY te.item_id
      ),
      open_po AS (
        SELECT poe.item_id, SUM(poe.quantity_ordered - poe.quantity_received) AS qty_open
        FROM purchase_order po JOIN purchase_order_entry poe ON poe.purchase_order_id = po.id
        WHERE po.status < 5 AND poe.quantity_ordered > poe.quantity_received
        GROUP BY poe.item_id
      )
      SELECT
        i.item_lookup_code AS "UPC",
        i.description AS "Description",
        COALESCE(d.name, '') AS "Department",
        COALESCE(c.name, '') AS "Category",
        COALESCE(s.supplier_name, '') AS "SupplierName",
        i.quantity AS "OnHand",
        COALESCE(op.qty_open, 0) AS "OpenPOQty",
        COALESCE(v.monthly_units, 0) AS "AvgMonthlySales",
        (CURRENT_DATE - i.last_sold::date)::int AS "DaysSinceLastSale",
        COALESCE(v.daily_units, 0) * 7 * COALESCE(i.price, 0) AS "EstLostSalesPerWeek",
        CASE
          WHEN COALESCE(v.daily_units,0) = 0 THEN 'Dead'
          WHEN COALESCE(op.qty_open, 0) = 0 THEN 'Critical'
          ELSE 'High'
        END AS "Risk",
        CASE
          WHEN COALESCE(op.qty_open, 0) > 0 THEN 'On PO — wait'
          WHEN COALESCE(v.daily_units, 0) = 0 THEN 'Delist'
          ELSE 'Reorder NOW'
        END AS "Action"
      FROM item i
      LEFT JOIN department d ON i.department_id = d.id
      LEFT JOIN category   c ON i.category_id   = c.id
      LEFT JOIN supplier   s ON i.supplier_id   = s.id
      LEFT JOIN velocity   v ON i.id = v.item_id
      LEFT JOIN open_po    op ON op.item_id = i.id
      WHERE {" AND ".join(conds)}
        AND COALESCE(v.daily_units, 0) > 0
        AND COALESCE(v.daily_units, 0) * 7 * COALESCE(i.price, 0) >= {float(min_lost)}
        {"AND op.qty_open > 0" if with_open_po == "yes" else ""}
        {"AND op.qty_open IS NULL" if with_open_po == "no" else ""}
      ORDER BY "EstLostSalesPerWeek" DESC NULLS LAST
      LIMIT {int(limit)}
    """
    rows = await fetch(sql, *args)
    lost = sum(float(r["EstLostSalesPerWeek"] or 0) for r in rows)
    on_po = sum(1 for r in rows if (r["OpenPOQty"] or 0) > 0)
    suppliers = len({r["SupplierName"] for r in rows if r["SupplierName"]})

    by_risk: dict[str, dict] = {}
    by_supp: dict[str, dict] = {}
    for r in rows:
        rk = r["Risk"]
        b = by_risk.setdefault(rk, {"Risk": rk, "Skus": 0, "LostSalesPerWeek": 0})
        b["Skus"] += 1
        b["LostSalesPerWeek"] += float(r["EstLostSalesPerWeek"] or 0)
        sn = r["SupplierName"] or "(none)"
        bs = by_supp.setdefault(sn, {"SupplierName": sn, "Skus": 0, "LostSalesPerWeek": 0})
        bs["Skus"] += 1
        bs["LostSalesPerWeek"] += float(r["EstLostSalesPerWeek"] or 0)

    return {
        "summary": {"sku_count": len(rows), "lost_sales_per_week": lost,
                    "open_po_covered": on_po, "suppliers_affected": suppliers},
        "by_risk":     sorted(by_risk.values(), key=lambda x: -x["LostSalesPerWeek"]),
        "by_supplier": sorted(by_supp.values(), key=lambda x: -x["LostSalesPerWeek"]),
        "rows": rows,
    }


@router.get("/stockouts/export.xlsx")
async def stockouts_xlsx(velocity_months: int = 6, supplier: Optional[str] = None,
                          dept: Optional[str] = None, min_lost: float = 0,
                          with_open_po: Optional[str] = None, limit: int = 10000,
                          _=Depends(_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    data = await stockouts(velocity_months=velocity_months, supplier=supplier,
                            dept=dept, min_lost=min_lost,
                            with_open_po=with_open_po, limit=limit)
    rows = data["rows"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Stockouts"
    if rows:
        cols = list(rows[0].keys())
        ws.append(cols)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFFFF")
            cell.fill = PatternFill("solid", fgColor="FF1B1F2B")
        for r in rows:
            ws.append([r.get(c) for c in cols])
    buf = io.BytesIO()
    wb.save(buf)
    return Response(content=buf.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": 'attachment; filename="stockouts.xlsx"'})


# --------------------------------------------------------------------- #
# Drill-down endpoints used by the React modal
# --------------------------------------------------------------------- #

@router.get("/item")
async def item_drill(id: str = Query(...), months: int = 24, _=Depends(_user)):
    """Look up an item by lookup-code (UPC) OR numeric id, then return
    the rich drill-down payload the modal expects."""
    by_code = await fetchrow("""
        SELECT i.id, i.item_lookup_code, i.description,
               COALESCE(d.name,'') AS department,
               COALESCE(c.name,'') AS category,
               i.quantity, i.quantity_committed, i.cost, i.price,
               i.last_sold, i.last_received
        FROM item i
        LEFT JOIN department d ON i.department_id = d.id
        LEFT JOIN category   c ON i.category_id   = c.id
        WHERE i.item_lookup_code = $1
           OR i.id::text = $1
        LIMIT 1
    """, id)
    if not by_code:
        return {"found": False, "identifier": id, "resolved_by": "exact-code+id"}
    item_id = by_code["id"]

    open_po = await fetchrow("""
        SELECT COALESCE(SUM(poe.quantity_ordered - poe.quantity_received), 0) AS qty_open
        FROM purchase_order po JOIN purchase_order_entry poe ON poe.purchase_order_id = po.id
        WHERE poe.item_id = $1 AND po.status < 5
          AND poe.quantity_ordered > poe.quantity_received
    """, item_id)

    velocity = await fetchrow("""
        SELECT SUM(te.quantity)/NULLIF($2,0) AS monthly,
               SUM(te.quantity)/NULLIF($2*30.0,0) AS daily
        FROM transaction_entry te
        WHERE te.item_id = $1
          AND te.transaction_time >= NOW() - make_interval(months => $2)
          AND te.quantity > 0
    """, item_id, max(1, months))

    monthly = float((velocity or {}).get("monthly") or 0)
    daily = float((velocity or {}).get("daily") or 0)
    on_hand = float(by_code["quantity"] or 0)
    committed = float(by_code["quantity_committed"] or 0)
    avail = on_hand - committed
    cur_mos = avail / monthly if monthly > 0 else None
    days_to_so = avail / daily if daily > 0 else None

    risk = (
        "Dead" if daily == 0 else
        "Excess" if (cur_mos or 0) > 12 else
        "Critical" if (cur_mos or 0) <= 0.25 else
        "High" if (cur_mos or 0) <= 0.7 else
        "Moderate" if (cur_mos or 0) <= 2 else
        "Healthy"
    )

    summary = await fetchrow("""
        SELECT COUNT(*)::int AS transaction_count,
               COALESCE(SUM(ABS(CASE WHEN te.quantity>0 THEN te.quantity END)),0) AS units_sold_abs,
               COALESCE(SUM(CASE WHEN te.quantity<0 THEN -te.quantity END),0) AS units_received,
               MAX(CASE WHEN te.quantity>0 THEN te.transaction_time END) AS last_sale,
               MAX(CASE WHEN te.quantity<0 THEN te.transaction_time END) AS last_receive
        FROM transaction_entry te
        WHERE te.item_id = $1
          AND te.transaction_time >= NOW() - make_interval(months => $2)
    """, item_id, max(1, months))

    txns = await fetch("""
        SELECT te.transaction_time AS "TxnDate",
               CASE WHEN te.quantity > 0 THEN 'SALE' ELSE 'PURCHASE/RECEIVE' END AS "TxnType",
               te.quantity AS "QtyImpact",
               te.price AS "UnitPrice",
               (te.quantity * te.price) AS "LineTotal",
               SUM(te.quantity) OVER (ORDER BY te.transaction_time
                                      ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW)
                 AS "RunningQty",
               'TXN#' || te.transaction_number::text AS "Reference"
        FROM transaction_entry te
        WHERE te.item_id = $1
          AND te.transaction_time >= NOW() - make_interval(months => $2)
        ORDER BY te.transaction_time
        LIMIT 500
    """, item_id, max(1, months))

    open_po_rows = await fetch("""
        SELECT po.po_number AS "PONumber",
               po.date_created AS "PODate",
               s.supplier_name AS "SupplierName",
               poe.quantity_ordered AS "QtyOrdered",
               poe.quantity_received AS "QtyReceived",
               (poe.quantity_ordered - poe.quantity_received) AS "QtyOpen",
               poe.price AS "UnitCost",
               (poe.quantity_ordered - poe.quantity_received) * poe.price AS "LineTotal"
        FROM purchase_order po
        JOIN purchase_order_entry poe ON poe.purchase_order_id = po.id
        LEFT JOIN supplier s ON po.supplier_id = s.id
        WHERE poe.item_id = $1 AND po.status < 5
          AND poe.quantity_ordered > poe.quantity_received
        ORDER BY po.date_created DESC
    """, item_id)

    return {
        "found": True,
        "history_months": months,
        "item": {
            "ID": item_id,
            "UPC": by_code["item_lookup_code"],
            "Description": by_code["description"],
            "Department": by_code["department"],
            "Category": by_code["category"],
            "OnHand": on_hand,
            "OpenPOQty": float((open_po or {}).get("qty_open") or 0),
            "AvgMonthlySales": monthly,
            "CurrentMoS": cur_mos,
            "DaysToStockout": days_to_so,
            "Cost": float(by_code["cost"] or 0),
            "Price": float(by_code["price"] or 0),
            "Risk": risk,
        },
        "summary": {
            "transaction_count": (summary or {}).get("transaction_count", 0),
            "units_sold_abs":   float((summary or {}).get("units_sold_abs") or 0),
            "units_received":   float((summary or {}).get("units_received") or 0),
            "last_sale_in_window":    (summary or {}).get("last_sale"),
            "last_receive_in_window": (summary or {}).get("last_receive"),
        },
        "transactions": txns,
        "open_pos": open_po_rows,
        "alt_matches": [],
    }


@router.get("/transaction")
async def txn_drill(ref: str = Query(...), _=Depends(_user)):
    """Resolve a TXN# or ITL# reference back to its full payload."""
    if not ref:
        return {"found": False, "reference": ref}
    if ref.upper().startswith("TXN#"):
        try:
            n = int(ref[4:])
        except ValueError:
            return {"found": False, "reference": ref}
        header = await fetchrow("""
            SELECT transaction_number AS "TransactionNumber",
                   time AS "TxnDate",
                   (total - sales_tax) AS "SubTotal",
                   sales_tax AS "SalesTax",
                   total AS "Total",
                   comment AS "Comment"
            FROM "transaction" WHERE transaction_number = $1
        """, n)
        if not header:
            return {"found": False, "reference": ref}
        lines = await fetch("""
            SELECT te.id AS "ID",
                   i.item_lookup_code AS "UPC",
                   i.description AS "Description",
                   COALESCE(d.name, '') AS "Department",
                   te.quantity AS "Quantity",
                   te.price AS "Price",
                   te.cost AS "Cost",
                   (te.quantity * te.price) AS "LineTotal"
            FROM transaction_entry te
            JOIN item i ON te.item_id = i.id
            LEFT JOIN department d ON i.department_id = d.id
            WHERE te.transaction_number = $1
            ORDER BY te.id
        """, n)
        total_qty = sum(float(l["Quantity"] or 0) for l in lines)
        total_val = sum(float(l["LineTotal"] or 0) for l in lines)
        return {"found": True, "kind": "transaction", "reference": ref,
                "type": "Sale", "event_time": str(header["TxnDate"]),
                "total_qty": total_qty, "total_value": total_val,
                "header": header, "lines": lines}
    return {"found": False, "reference": ref, "error": "Unsupported reference type."}
